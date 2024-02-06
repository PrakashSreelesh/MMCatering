from datetime import datetime
from tkinter import ttk
import tkinter as tk
from tkinter import *
from datetime import date, datetime
from tkinter import filedialog
from tkinter import messagebox
from tkcalendar import DateEntry
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl
import xlrd
from openpyxl import Workbook
import pathlib


def on_date_selected():
    selected_date = startDateEntry.get_date()
    StartDate.set(selected_date)


def search():
    # Retrieve the values from the entry widgets
    search_name = Name.get()
    search_mobile = Mobile.get()
    search_start_date = startDateEntry.get_date()

    formatted_start_date = search_start_date.strftime('%Y-%m-%d')
    print(search_name, search_mobile, formatted_start_date)

    # Load the workbook
    workbook = openpyxl.load_workbook('MMUsersList.xlsx')
    sheet = workbook.active

    result = ""

    # Iterate through rows and search for the criteria
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if (search_name and search_name.lower() in row[2].lower()) or \
           (search_mobile and int(search_mobile) == row[3]) or \
           (search_start_date and formatted_start_date == row[5].strftime('%Y-%m-%d')):
            result += f"Reg No: {row[0]}, Name: {row[2]}, Contact No: {row[3]}, Start Date: {row[5]}\n"

    # Update the result label
    resultLabel.config(text=result)

    # Print the result in the terminal
    print(result)

    # Close the workbook
    workbook.close()


# Get the current date
current_date = datetime.today().date()

bgClr = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

root = Tk()
root.title("MM Catering User Management")
root.geometry("1250x700+210+100")
root.config(bg=bgClr)


file = pathlib.Path('MMUsersList.xlsx')

if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Reg No"
    sheet['B1'] = "Building No"
    sheet['C1'] = "Name"
    sheet['D1'] = "Contact No"
    sheet['E1'] = "Address"
    sheet['F1'] = "Start Date"
    sheet['G1'] = "End Date"
    sheet['H1'] = "Category"
    sheet['I1'] = "Fees"
    sheet['J1'] = "Total Days"
    sheet['K1'] = "Payment Date"
    sheet['L1'] = "Advance"
    sheet['M1'] = "Prv Due Amount"
    sheet['N1'] = "Extra"
    sheet['O1'] = "Total Amount"
    sheet['P1'] = "Paid Amount"
    sheet['Q1'] = "Balance to Pay"

    file.save('MMUsersList.xlsx')


# Search box to update
Label(root, text="Name :", font="arial 13",
      fg=framebg, bg=bgClr).place(x=30, y=30)
Label(root, text="Mob :", font="arial 13",
      fg=framebg, bg=bgClr).place(x=310, y=30)
Label(root, text="Start Date:", font="arial 13",
      fg=framebg, bg=bgClr).place(x=560, y=30)
Label(root, text="End Date :", font="arial 13",
      fg=framebg, bg=bgClr).place(x=830, y=30)

Name = StringVar()
Mobile = StringVar()
StartDate = tk.StringVar()
EndDate = StringVar()

nameEntry = Entry(root, textvariable=Name, width=25, font="arial 10")
mobEntry = Entry(root, textvariable=Mobile, width=24, font="arial 10")
startDateEntry = DateEntry(root, width=12, background='darkblue', foreground='white', borderwidth=2,
                           year=current_date.year, month=current_date.month, day=current_date.day)
endDateEntry = Entry(root, textvariable=EndDate, width=20, font="arial 10")

nameEntry.place(x=100, y=30)
mobEntry.place(x=370, y=30)
startDateEntry.place(x=660, y=30)
endDateEntry.place(x=920, y=30)

SearchBtn = Button(root, text="Search", command=search, width=10,
                   bg='#68ddfa', font="arial 13 bold")
SearchBtn.place(x=1100, y=25)

resultLabel = Label(root, text="", font="arial 12", bg=bgClr, fg="green")
resultLabel.place(x=30, y=70)

root.mainloop()
